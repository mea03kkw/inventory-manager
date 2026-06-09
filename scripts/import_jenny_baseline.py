"""
Import Jenny's 231-row baseline from SampleLibrary_Data.xlsx.

Resets the inventory table, seeds master data (if not already seeded),
normalizes values, generates sample_code/serial_num, and validates.

Usage:
    python scripts/import_jenny_baseline.py [--force]
"""

import os
import sys
import argparse
from urllib.parse import urlparse
import psycopg2
import psycopg2.extensions
from datetime import date, datetime
from dotenv import load_dotenv

load_dotenv()

parser = argparse.ArgumentParser()
parser.add_argument("--force", action="store_true", help="Skip confirmation prompt")
parser.add_argument("--allow-remote", action="store_true", help="Allow connecting to remote/non-localhost databases")
parser.add_argument("--db-url", type=str, default=None, help="Database URL (overrides env and .env)")
args = parser.parse_args()

DATABASE_URL = ""
if args.db_url:
    DATABASE_URL = args.db_url
else:
    load_dotenv()
    DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
if not DATABASE_URL:
    sys.exit("ERROR: DATABASE_URL is not set")

# Safety guard: reject remote hosts unless --allow-remote is passed
_parsed = urlparse(DATABASE_URL)
_is_remote = _parsed.hostname and _parsed.hostname not in ("localhost", "127.0.0.1", "::1", "db", "postgres")
if _is_remote and not args.allow_remote:
    sys.exit(
        f"ERROR: Refusing to connect to remote host '{_parsed.hostname}' without --allow-remote flag.\n"
        f"This script runs TRUNCATE ... CASCADE which is destructive.\n"
        f"Pass --allow-remote only if you are certain this is the correct database."
    )

# Category normalization map
CATEGORY_MAP = {
    "dryer": "Dryer",
    "styler": "Styler",
    "style": "Styler",
    "straightener": "Straightener",
    "brush & massager": "Brush & Massager",
    "others": "Others",
}

# ============================================================
# Helpers
# ============================================================

def normalize_brand(raw):
    v = (raw or "").strip()
    if v.lower() == "philips":
        return "Philips"
    return v

def derive_sample_type(brand):
    return "Philips" if normalize_brand(brand).lower() == "philips" else "Competitor"

def derive_prefix(sample_type):
    return "PHI" if sample_type == "Philips" else "CMT"

def normalize_category(raw):
    v = (raw or "").strip().lower()
    return CATEGORY_MAP.get(v, v.title() if v else "")

def extract_year(dt_val):
    if dt_val is None:
        return ""
    if isinstance(dt_val, datetime):
        return str(dt_val.year)
    if isinstance(dt_val, date):
        return str(dt_val.year)
    try:
        return str(date.fromisoformat(str(dt_val).split()[0]).year)
    except (ValueError, TypeError):
        return ""

def generate_sample_code(prefix, year, db_id):
    return f"{prefix}{year}-{db_id:04d}"

def derive_serial_num(sample_code):
    return sample_code[3:].replace("-", "")

# ============================================================
# Main
# ============================================================

def main():
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()

    # --- Step 0: Backup check ---
    cur.execute("SELECT COUNT(*) FROM inventory")
    existing_count = cur.fetchone()[0]
    print(f"Current inventory row count: {existing_count}")
    if existing_count > 0 and not args.force:
        print("WARNING: Inventory is not empty. Data will be cleared.")
        confirm = input("Type 'RESET' to proceed: ")
        if confirm != "RESET":
            print("Aborted.")
            conn.close()
            return

    # --- Step 1: Clear inventory (outside transaction) ---
    print("Clearing inventory and checkout records...")
    conn.set_isolation_level(psycopg2.extensions.ISOLATION_LEVEL_AUTOCOMMIT)
    cur.execute("TRUNCATE TABLE checkout_records, inventory RESTART IDENTITY CASCADE")
    print("Tables truncated.")
    # Return to default isolation for the transactional insert
    conn.set_isolation_level(psycopg2.extensions.ISOLATION_LEVEL_READ_COMMITTED)

    # --- Step 2: Read Excel ---
    import openpyxl
    wb = openpyxl.load_workbook(
        os.path.join(os.path.dirname(os.path.dirname(__file__)), "SampleLibrary_Data.xlsx"),
        data_only=True,
    )
    ws = wb.active
    rows_data = []
    for row_idx in range(2, ws.max_row + 1):
        row = {}
        row["SampleCode"] = ws.cell(row=row_idx, column=1).value or ""
        row["SerialNum"] = ws.cell(row=row_idx, column=2).value or ""
        row["SampleType"] = ws.cell(row=row_idx, column=3).value or ""
        row["ProductName"] = ws.cell(row=row_idx, column=4).value or ""
        row["Brand"] = ws.cell(row=row_idx, column=5).value or ""
        row["Model"] = ws.cell(row=row_idx, column=6).value or ""
        row["Category"] = ws.cell(row=row_idx, column=7).value or ""
        row["SubCategory"] = ws.cell(row=row_idx, column=8).value or ""
        row["DepartmentOwner"] = ws.cell(row=row_idx, column=9).value or ""
        row["Condition"] = ws.cell(row=row_idx, column=10).value or ""
        row["DateReceived"] = ws.cell(row=row_idx, column=11).value
        row["StorageLocationCode"] = ws.cell(row=row_idx, column=12).value or ""
        row["Environment"] = ws.cell(row=row_idx, column=13).value or ""
        row["UnitCount"] = ws.cell(row=row_idx, column=14).value or ""
        row["Remark"] = ws.cell(row=row_idx, column=15).value or ""
        rows_data.append(row)

    print(f"Read {len(rows_data)} rows from Excel")

    # --- Step 4: Normalize ---
    print("Normalizing data...")
    for i, row in enumerate(rows_data):
        db_id = i + 1

        brand = normalize_brand(row["Brand"])
        sample_type = derive_sample_type(brand)
        prefix = derive_prefix(sample_type)
        year = extract_year(row["DateReceived"])
        sample_code = generate_sample_code(prefix, year, db_id)
        serial_num = derive_serial_num(sample_code)
        category = normalize_category(row["Category"])

        # Handle null DateReceived
        date_received_str = ""
        if row["DateReceived"] is not None:
            if isinstance(row["DateReceived"], (datetime, date)):
                date_received_str = row["DateReceived"].strftime("%Y-%m-%d")
            else:
                date_received_str = str(row["DateReceived"]).strip()

        # Handle legacy blank text fields: fill with NA
        def na_if_blank(v):
            if v is None:
                return "NA"
            v = str(v).strip()
            return v if v else "NA"

        model = na_if_blank(row["Model"])
        sub_category = na_if_blank(row["SubCategory"])
        condition = na_if_blank(row["Condition"])
        remark = na_if_blank(row["Remark"])
        product_name = na_if_blank(row["ProductName"])

        # DepartmentOwner — normalize dept code
        dept = (row["DepartmentOwner"] or "").strip()
        dept_code_map = {
            "r&d": "RD",
            "prc": "PRC",
            "cmm": "CMM",
            "pmo": "PMO",
            "others": "OTH",
        }
        dept_code = dept_code_map.get(dept.lower().strip(), dept)

        # UnitCount
        unit_count = row["UnitCount"]
        try:
            uc = int(str(unit_count).strip()) if unit_count else 1
            uc = uc if uc > 0 else 1
        except (ValueError, TypeError):
            uc = 1

        # Environment — use "Legacy" as default if blank
        environment = (row["Environment"] or "").strip()
        if not environment:
            environment = "Legacy"

        rows_data[i] = {
            "db_id": db_id,
            "Title": na_if_blank(row["ProductName"]),
            "SerialNum": serial_num,
            "SampleType": sample_type,
            "ProductName": product_name,
            "Brand": brand,
            "Model": model,
            "Category": category,
            "SubCategory": sub_category,
            "DepartmentOwner": dept_code,
            "Condition": condition,
            "DateReceived": date_received_str,
            "StorageLocationCode": (row["StorageLocationCode"] or "").strip(),
            "UnitCount": str(uc),
            "UnitMeasure": "",
            "Column1": "",
            "Attachments": "",
            "Notes": remark,
            "PhotoLink": "",
            "Status": "IN_STOCK",
            "sample_code": sample_code,
            "record_state": "ACTIVE",
            "Environment": environment,
        }

    # --- Step 5: Insert ---
    print("Inserting 231 rows...")
    all_fields = [
        "Title", "SerialNum", "SampleType", "ProductName", "Brand", "Model",
        "Category", "SubCategory", "DepartmentOwner", "Condition", "DateReceived",
        "StorageLocationCode", "UnitCount", "UnitMeasure", "Column1", "Attachments",
        "Notes", "PhotoLink",
    ]
    new_fields = ["sample_code", "record_state", "Environment"]
    all_fields_v2 = all_fields + new_fields

    placeholders = ", ".join(["%s"] * len(all_fields_v2))
    field_sql = ", ".join([f'"{f}"' for f in all_fields_v2])

    insert_sql = f"INSERT INTO inventory ({field_sql}, \"Status\", quantity, available_quantity) VALUES ({placeholders}, %s, %s, %s)"

    conn.autocommit = False
    try:
        for r in rows_data:
            values = [r.get(f) for f in all_fields_v2]
            values.append(r["Status"])
            values.append(int(r["UnitCount"]))
            values.append(int(r["UnitCount"]))
            cur.execute(insert_sql, values)
        conn.commit()
        print("Insert successful!")
        # Ensure sequence is exactly at MAX(id) for the next insert
        cur.execute("SELECT setval('inventory_id_seq', (SELECT MAX(id) FROM inventory))")
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        sys.exit(f"ERROR during insert: {e}")

    # --- Step 6: Verify ---
    print("\n=== VERIFICATION ===")
    cur.execute("SELECT COUNT(*) FROM inventory")
    count = cur.fetchone()[0]
    print(f"Total rows: {count}")
    assert count == 231, f"Expected 231, got {count}"

    cur.execute("SELECT id, sample_code, \"SerialNum\", \"SampleType\", \"Brand\", \"DateReceived\" FROM inventory ORDER BY id")
    rows = cur.fetchall()

    # Check sequence
    first_id = rows[0][0]
    last_id = rows[-1][0]
    print(f"First ID: {first_id}, Last ID: {last_id}")
    assert first_id == 1, f"First ID should be 1, got {first_id}"
    assert last_id == 231, f"Last ID should be 231, got {last_id}"

    # Check codes
    errors = []
    for r in rows:
        rid, scode, sernum, stype, brand, drecv = r
        expected_type = "Philips" if (brand or "").strip().lower() == "philips" else "Competitor"
        prefix = "PHI" if expected_type == "Philips" else "CMT"
        year = extract_year(drecv)
        expected_code = generate_sample_code(prefix, year, rid)
        expected_serial = derive_serial_num(expected_code)

        if scode != expected_code:
            errors.append(f"ID {rid}: SampleCode mismatch: got '{scode}', expected '{expected_code}'")
        if sernum != expected_serial:
            errors.append(f"ID {rid}: SerialNum mismatch: got '{sernum}', expected '{expected_serial}'")
        if stype != expected_type:
            errors.append(f"ID {rid}: SampleType mismatch: got '{stype}', expected '{expected_type}'")

    # Check uniqueness
    cur.execute("SELECT COUNT(*), COUNT(DISTINCT sample_code), COUNT(DISTINCT \"SerialNum\") FROM inventory")
    total, unique_sc, unique_sn = cur.fetchone()
    if unique_sc != 231:
        errors.append(f"Duplicate sample_code found: {unique_sc} unique out of {total}")
    if unique_sn != 231:
        errors.append(f"Duplicate SerialNum found: {unique_sn} unique out of {total}")

    if errors:
        print("\n!!! VALIDATION ERRORS:")
        for e in errors[:20]:
            print(f"  {e}")
        if len(errors) > 20:
            print(f"  ... and {len(errors) - 20} more")
    else:
        print("All validations PASSED!")
        print("  - 231 rows imported")
        print("  - IDs: 1 to 231")
        print("  - SampleCode unique")
        print("  - SerialNum unique")
        print("  - SampleType matches Brand rule")
        print("  - SampleCode year matches DateReceived year")

    # Check next sequence value
    cur.execute("SELECT nextval('inventory_id_seq')")
    next_id = cur.fetchone()[0]
    print(f"\nNext sequence value: {next_id}")
    assert next_id == 232, f"Expected next ID = 232, got {next_id}"

    conn.close()
    print("\n=== IMPORT COMPLETE ===")

if __name__ == "__main__":
    main()